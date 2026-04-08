import streamlit as st
import pandas as pd
from lxml import etree
from io import BytesIO
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import unicodedata

# ===============================
# YARDIMCI FONKSİYONLAR
# ===============================
def clean(val):
    if pd.isna(val) or str(val).strip() == "":
        return None
    v = str(val).strip()
    v = unicodedata.normalize("NFKC", v)
    return v

def safe_int_str(val):
    if pd.isna(val) or str(val).strip() == "":
        return None
    try:
        return str(int(float(val)))
    except:
        return str(val)

def autofit_columns(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    val_len = len(str(cell.value))
                    if val_len > max_length:
                        max_length = val_len
            except:
                pass
        adjusted_width = max(15, min(max_length + 5, 50))
        ws.column_dimensions[column_letter].width = adjusted_width

def apply_modern_style(ws):
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    thin_side = Side(style='thin', color="D6DBDF")
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border
    
    ws.freeze_panes = "A2"
    autofit_columns(ws)

# ===============================
# XSD ANALİZ SİSTEMİ
# ===============================
def get_xsd_details(xsd_file):
    try:
        parser = etree.XMLParser(remove_blank_text=True)
        tree = etree.parse(xsd_file, parser)
        root = tree.getroot()
        
        def fetch_enums(search_term):
            path1 = f".//{{*}}simpleType[@name='{search_term}']//{{*}}enumeration"
            enums = [el.get("value") for el in root.findall(path1)]
            if not enums:
                element = root.find(f".//{{*}}element[@name='{search_term}']")
                if element is not None:
                    type_attr = element.get("type")
                    if type_attr:
                        clean_type = type_attr.split(':')[-1]
                        path2 = f".//{{*}}simpleType[@name='{clean_type}']//{{*}}enumeration"
                        enums = [el.get("value") for el in root.findall(path2)]
            return enums

        return {
            "SerbestBolgeAdi": fetch_enums("serbestBolgeType") or fetch_enums("SerbestBolgeAdi"),
            "FirmaFaaliyetRuhsatiKonusu": fetch_enums("ruhsatKonulariType") or fetch_enums("FirmaFaaliyetRuhsatiKonusu"),
            "OlcuBirimleri": fetch_enums("olcuBirimiType") or fetch_enums("OlcuBirimleri"),
            "Ulkeler": fetch_enums("ulkeType") or fetch_enums("Ulkeler"),
            "ReferansFormTipi": fetch_enums("formTipiType") or fetch_enums("ReferansFormTipi")
        }
    except Exception as e:
        st.error(f"XSD Okuma Hatası: {e}")
        return {}

# ===============================
# STREAMLIT ARAYÜZÜ
# ===============================
st.set_page_config(page_title="XSD to XML Converter", layout="wide")
st.title("🚀 XSD Validation Excel to XML Converter")

app_mode = st.sidebar.radio("İşlem Menüsü:", ["1. Şablon Oluştur (XSD)", "2. XML'e Dönüştür (Excel)"])

if app_mode == "1. Şablon Oluştur (XSD)":
    st.header("📂 1. Adım: XSD Yükle ve Şablon Hazırla")
    uploaded_xsd = st.file_uploader("Sistemin XSD dosyasını yükleyin", type=["xsd"])

    if uploaded_xsd:
        xsd_data = get_xsd_details(uploaded_xsd)
        output = BytesIO()
        wb = Workbook()
        
        # --- KULLANIM KILAVUZU ---
        ws_guide = wb.active
        ws_guide.title = "Kullanim Kilavuzu"
        guide_data = [
            ["XSD VALIDATION EXCEL TO XML CONVERTER"],
            [""],
            ["NASIL KULLANILIR?"],
            ["1. GenelBilgiler", "İlgili alanları doldurun."],
            ["2. Urunler", "Her ürüne benzersiz bir 'UrunSiraNo' verin (Örn: 1, 2, 3)."],
            ["3. Hammaddeler", "Hammaddenin hangi ürüne ait olduğunu 'BagliUrunSiraNo' sütununa yazın."],
            ["4. Miktarlar", "Birinci, İkinci ve Üçüncü miktar/birim alanlarını ihtiyaca göre doldurun."],
            ["5. Fireler", "Fire miktarı varsa ilgili sütunlara (Örn: BirinciFireMiktar) girin."],
            [""],
            ["⚠️ DİKKAT:", "Sütun başlıklarını değiştirmeyin. Sadece listedeki (Dropdown) değerleri kullanın."]
        ]
        for row in guide_data:
            ws_guide.append(row)
        ws_guide["A1"].font = Font(size=14, bold=True, color="2C3E50")
        ws_guide.column_dimensions['A'].width = 25
        ws_guide.column_dimensions['B'].width = 65

        # --- GENEL BİLGİLER ---
        ws_g = wb.create_sheet("GenelBilgiler")
        ws_g.append(["DisReferansNo", "SerbestBolgeAdi", "FirmaFaaliyetRuhsatiNo", "FirmaFaaliyetRuhsatiKonusu", "GirisTarihi"])

        # --- ÜRÜNLER (Gelişmiş) ---
        ws_u = wb.create_sheet("Urunler")
        ws_u.append(["UrunSiraNo", "gtip", "UrunAdi", "BirinciMiktar", "BirinciBirim", 
                     "IkinciMiktar", "IkinciBirim", "UcuncuMiktar", "UcuncuBirim", "UrunMensei"])

        # --- HAMMADDELER (TAM EKSİKSİZ LİSTE) ---
        ws_h = wb.create_sheet("Hammaddeler")
        ws_h.append([
            "BagliUrunSiraNo", "ReferansFormTipi", "ReferansFormNo", "ReferansFormYil", "ReferansSiraNo",
            "gtip", "Cins", "Mensei", 
            "BirinciMiktar", "BirinciFireMiktar", "BirinciBirim",
            "IkinciMiktar", "IkinciFireMiktar", "IkinciBirim",
            "UcuncuMiktar", "UcuncuFireMiktar", "UcuncuBirim"
        ])

        # --- LİSTELER VE DOĞRULAMA ---
        ws_l = wb.create_sheet("Listeler")
        col_map = {"A": "SerbestBolgeAdi", "B": "FirmaFaaliyetRuhsatiKonusu", "C": "OlcuBirimleri", "D": "Ulkeler", "E": "ReferansFormTipi"}
        for col, key in col_map.items():
            items = list(set(xsd_data.get(key, [])))
            for i, val in enumerate(items, 1):
                ws_l[f"{col}{i}"] = val

        def add_dv(ws, list_col, list_size, range_str):
            if list_size > 0:
                dv = DataValidation(type="list", formula1=f"Listeler!${list_col}$1:${list_col}${list_size}", allow_blank=True)
                ws.add_data_validation(dv)
                dv.add(range_str)

        # Doğrulamaları uygula
        add_dv(ws_g, "A", len(xsd_data.get("SerbestBolgeAdi", [])), "B2:B1000")
        add_dv(ws_u, "C", len(xsd_data.get("OlcuBirimleri", [])), "E2:E1000")
        add_dv(ws_u, "C", len(xsd_data.get("OlcuBirimleri", [])), "G2:G1000")
        add_dv(ws_u, "C", len(xsd_data.get("OlcuBirimleri", [])), "I2:I1000")
        add_dv(ws_u, "D", len(xsd_data.get("Ulkeler", [])), "J2:J1000")
        
        add_dv(ws_h, "E", len(xsd_data.get("ReferansFormTipi", [])), "B2:B5000")
        add_dv(ws_h, "D", len(xsd_data.get("Ulkeler", [])), "H2:H5000")
        add_dv(ws_h, "C", len(xsd_data.get("OlcuBirimleri", [])), "K2:K5000")
        add_dv(ws_h, "C", len(xsd_data.get("OlcuBirimleri", [])), "N2:N5000")
        add_dv(ws_h, "C", len(xsd_data.get("OlcuBirimleri", [])), "Q2:Q5000")

        for sheet in wb.worksheets:
            if sheet.title != "Listeler":
                apply_modern_style(sheet)
        
        ws_l.sheet_state = "hidden"
        wb.save(output)
        st.success("Tüm sütunları içeren şablon hazırlandı!")
        st.download_button("📥 Tam Şablonu İndir", output.getvalue(), "UBF_Full_Sablon.xlsx")

elif app_mode == "2. XML'e Dönüştür (Excel)":
    st.header("🛠️ 2. Adım: Verileri XML'e Dönüştür")
    excel_file = st.file_uploader("Doldurulmuş Excel dosyasını yükleyin", type=["xlsx"])

    if excel_file:
        try:
            df_g = pd.read_excel(excel_file, sheet_name="GenelBilgiler")
            df_u = pd.read_excel(excel_file, sheet_name="Urunler")
            df_h = pd.read_excel(excel_file, sheet_name="Hammaddeler")

            root = etree.Element("UBFBilgileri")

            # --- GENEL BİLGİLER ---
            if not df_g.empty:
                g = df_g.iloc[0]
                genel = etree.SubElement(root, "UBFGenelBilgiler")
                for f in ["DisReferansNo", "SerbestBolgeAdi", "FirmaFaaliyetRuhsatiNo", "FirmaFaaliyetRuhsatiKonusu", "GirisTarihi"]:
                    val = clean(g.get(f))
                    if val: etree.SubElement(genel, f).text = val

            # --- ÜRÜN VE HAMMADDE DÖNGÜSÜ ---
            for _, u in df_u.iterrows():
                u_el = etree.SubElement(root, "UrunBilgileri")
                urun = etree.SubElement(u_el, "Urun")
                
                gt_u = "".join(filter(str.isdigit, str(clean(u.get("gtip")) or "")))
                etree.SubElement(urun, "gtip").text = gt_u.zfill(12)[:12]
                etree.SubElement(urun, "UrunAdi").text = clean(u.get("UrunAdi")) or ""
                etree.SubElement(urun, "BirinciMiktar").text = str(u.get("BirinciMiktar") or "0")
                etree.SubElement(urun, "BirinciBirim").text = clean(u.get("BirinciBirim")) or ""
                
                # Opsiyonel 2. ve 3. Miktarlar
                if clean(u.get("IkinciMiktar")):
                    etree.SubElement(urun, "IkinciMiktar").text = str(u.get("IkinciMiktar"))
                    etree.SubElement(urun, "IkinciBirim").text = clean(u.get("IkinciBirim")) or ""
                if clean(u.get("UcuncuMiktar")):
                    etree.SubElement(urun, "UcuncuMiktar").text = str(u.get("UcuncuMiktar"))
                    etree.SubElement(urun, "UcuncuBirim").text = clean(u.get("UcuncuBirim")) or ""
                
                etree.SubElement(urun, "UrunMensei").text = clean(u.get("UrunMensei")) or ""

                u_id = safe_int_str(u.get("UrunSiraNo"))
                if u_id:
                    bagli_hamlar = df_h[df_h["BagliUrunSiraNo"].apply(safe_int_str) == u_id]
                    for _, h in bagli_hamlar.iterrows():
                        ham = etree.SubElement(u_el, "HamMadde")
                        etree.SubElement(ham, "ReferansFormTipi").text = clean(h.get("ReferansFormTipi")) or ""
                        etree.SubElement(ham, "ReferansFormNo").text = clean(h.get("ReferansFormNo")) or ""
                        etree.SubElement(ham, "ReferansFormYil").text = safe_int_str(h.get("ReferansFormYil")) or ""
                        
                        gt_h = "".join(filter(str.isdigit, str(clean(h.get("gtip")) or "")))
                        etree.SubElement(ham, "gtip").text = gt_h.zfill(12)[:12]
                        etree.SubElement(ham, "Mensei").text = clean(h.get("Mensei")) or ""
                        
                        # Hammadde 3'lü Birim Yapısı (Miktar + Fire + Birim)
                        birim_gruplari = [
                            ("BirinciMiktar", "BirinciFireMiktar", "BirinciBirim"),
                            ("IkinciMiktar", "IkinciFireMiktar", "IkinciBirim"),
                            ("UcuncuMiktar", "UcuncuFireMiktar", "UcuncuBirim")
                        ]
                        
                        for m, f, b in birim_gruplari:
                            if clean(h.get(m)):
                                etree.SubElement(ham, m).text = str(h.get(m))
                                if clean(h.get(f)):
                                    etree.SubElement(ham, f).text = str(h.get(f))
                                etree.SubElement(ham, b).text = clean(h.get(b)) or ""

            xml_data = etree.tostring(root, pretty_print=True, xml_declaration=True, encoding="UTF-8")
            st.success("Tüm veriler başarıyla XML'e dönüştürüldü!")
            st.download_button("✅ XML'i İndir", xml_data, "UBF_Full_Cikti.xml", "application/xml")
        except Exception as e:
            st.error(f"Dönüştürme Hatası: {e}")
